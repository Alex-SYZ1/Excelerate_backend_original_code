import os,sys
import openpyxl as px
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter

"""用于导入项目中不在同一文件夹的库"""
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import utils.excel_processor as XPRO

class FileRuleMaker:
    def __init__(self):
        pass
    def extract_fields_from_excel(self, excel_got):
        """
        接收Excel文件，读取第一个worksheet，寻找字段所在行，并提取所有字段值到列表。
        将字段行高亮，返回含字段名的列表和保存的Excel文件路径。

        Parameters:
        excel_got (file): 需要处理的Excel文件。

        Returns:
            ①
            content:list_saved_in_json_stream: 字段名的列表
            format :['序号', '作品题目', '参赛类别', '作品学科分类']
            ②
            content:excel_file_in_json_stream:需要提取字段的Excel文件。
            format : 
            
            
        """
        #获取文件并转化
        Xio=XPRO.Excel_IO()
        ""#项目实际部署时，无需判断是否为字符串，全部为前端发送的数据流 即改为excel_wb,excel_ws=Xio.load_workbook_from_stream(excel_got)
        excel_wb,excel_ws=Xio.load_workbook_from_stream(excel_got) if type(excel_got) !=str else Xio.read_excel_file(excel_got)
        ""
        #读取对象并获取属性
        Xattr=XPRO.Excel_attribute(excel_wb,excel_ws)
        All_rows=[Xattr.get_some_axis_cells(index,value_only=False) for index in range(1,min(Xattr.get_max_row_col()["max_row"])+1)]
        
        #对预知的字段行进行操作
        Predict_fields_cells   = max(All_rows,key=len)
        Predict_fields_values  = [cell.value for cell in Predict_fields_cells]
        Predict_fields_indexes = [cell.coordinate for cell in Predict_fields_cells]
        Xattr.modify_MutipleRange_style(Predict_fields_indexes,fill=PatternFill(fill_type='solid', start_color='FFFF00'))
        
        ""#演示用，避免数据流无法理解,可去除
        excel_wb.save("../tests/for_fuker.extract/output_test1.xlsx")
        ""
        OUTPUT1=XPRO.convert_to_json_stream(Predict_fields_values)
        OUTPUT2=Xio.stream_excel_to_frontend(excel_wb)
    
        return (OUTPUT1,OUTPUT2)
    
    def generate_user_rule_dict(self, excel_got, fields_list,field_row_num=1):
        """字段名所在行改为后端用函数获取高亮所在行号
        接收确认了字段行的Excel和字段名列表，输出包含预定义规则和下拉列表的字典到JSON。

        Parameters:
        excel_got (file): 确认了字段行的Excel文件。
        fields_list (list): 字段名列表。
        field_row_num (int):字段名所在行号,测试时默认设为1
        Returns:[均为转化后]
            content:dict_saved_in_json_stream: 用户可选规则字典。
            format :{"字段名1":{"对应列下拉列表规则":[["下拉规则1选项1","下拉规则1选项2"],
                                                    ["下拉规则2选项1","下拉规则2选项2"]],
                                "程序预定义规则":   [["程序预定义规则1选项1","程序预定义规则1选项2"],
                                                    ["程序预定义规则2选项1","程序预定义规则2选项2"]]}
                     "字段名2":...同上}
        """
        pass  
        # TODO: 
        # 将列号与单元格、字段名匹配  
        # 设定用户可选规则字典 
        # 匹配字段名与预定义规则(调用函数) 
        # 匹配下拉列表信息(调用函数) 
        # 返回用户可选规则字典
        
        # 复制自第一个方法
        ""
        # 获取文件并转化
        Xio=XPRO.Excel_IO()
        """项目实际部署时，无需判断是否为字符串，全部为前端发送的数据流"""
        excel_wb,excel_ws=Xio.load_workbook_from_stream(excel_got) if type(excel_got) !=str else Xio.read_excel_file(excel_got)
        
        # 读取对象并获取属性
        Xattr=XPRO.Excel_attribute(excel_wb,excel_ws)
        ""
        Field_row=Xattr.get_some_axis_cells(field_row_num,value_only=False) 
        #return Field_row
        # 匹配字段列号与字段单元格、列表中字段名，循环结束时应得到一个字段齐全的字典
        Field_index_to_cell_name={}
        for cell in Field_row:
            for name in fields_list:
                #print(name,cell.value,name==cell.value)
                if name==cell.value:
                    Field_index_to_cell_name[(get_column_letter(cell.column))]=[cell,name]
                    continue
        #return Field_index_to_cell_name
        # 设定用户可选规则字典 注：Python 3.6之后，字典是有序的
        Sheet_dropdowns=Xattr.get_dropdowns()
        Field_rules={name: dict(zip(["对应列下拉列表规则","程序预定义规则"],[Sheet_dropdowns[col_index] if col_index in Sheet_dropdowns else [],["syz随便写的程序预设规则1","syz随便写的程序预设规则2"]])) for col_index,(cell,name) in Field_index_to_cell_name.items() } 
        for j,k in Field_rules.items():
            print("*",j,k)
            
        OUTPUT=XPRO.convert_to_json_stream(Field_rules)
        return OUTPUT


    def create_final_rules_and_examples(self, rules_json):
        """
        接收前端传来的规则字典所在的JSON文件，生成最终规则和样例。

        Parameters:
        rules_json (json): 规则字典所在的JSON文件。

        Returns:
        dict: 包含规则和样例的JSON对象。
        """
        pass  # TODO: 实现方法

    def save_final_rules(self, rules_dict, additional_info):
        """
        接收最终确认后的规则字典和其他提示信息，保存为JSON文件。

        Parameters:
        rules_dict (dict): 规则字典。
        additional_info (str): 其他提示信息。

        Returns:
        str: 保存的JSON文件路径。
        """
        pass  # TODO: 实现方法
    
if "__main__" == __name__:
    
    Fuker=FileRuleMaker()
    excel_got=r"../tests/for_fuker.extract/test1.xlsx"
    # 测试第一个方法
    """
    print(Fuker.extract_fields_from_excel(excel_got))"""
    
    # 测试第二个方法
    fields_list=['序号', '作品题目', '参赛类别', '作品学科分类', '学科门类', '一级学科', '作者', '是否为团队负责人', '性别', '生源地', '学号', '所在院系', '年级（如2020级本科生/硕士生/博士生）', '手机', '微信号', '邮箱', '指导教师姓名', '指导教师性别', '指导教师所在院系', '指导教师职称/职务', '指导教师电话', '指导教师电子邮箱']
    excel_got=r"..\tests\for_fuker.extract\test2_dropdown_hidensheet.xlsx"
    (Fuker.generate_user_rule_dict(excel_got,fields_list,field_row_num=5))
