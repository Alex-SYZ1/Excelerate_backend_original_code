import openpyxl as px
import io,json,re,os
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import range_boundaries
import win32com.client as win32


class Excel_IO:
    def __init__(self):
        # Excel格式映射
        self.FORMATS = {'xls': 56,    'xlsx': 51}
        self.temp_path="tmp/"
    def read_excel_file(self, excel_path, sheet_index=0):
        """openpyxl读取某路径的excel文件"""
        try:
            excel_wb = px.load_workbook(excel_path, data_only=True)
            excel_ws = excel_wb.worksheets[sheet_index]
            return (excel_wb,excel_ws)
        except IOError as e:
            print(f"An error occurred during reading: {e}")
            # Handle the exception as needed
            return None

    def load_workbook_from_stream(self,excel_stream, sheet_index=0):
        """openpyxl读取某数据流的excel文件"""
        if 1:#try:
            # 读取流中的内容为二进制数据
            excel_data = excel_stream.read()
            # 使用BytesIO创建一个类似文件的对象
            excel_bytes = io.BytesIO(excel_data)
            return self.read_excel_file(excel_bytes)


    def save_excel(self, excel_wb, excel_path):
        """openpyxl传输wb对象到excel文件"""
        try:
            excel_wb.save(excel_path)
        except IOError as e:
            print(f"An error occurred during saving: {e}")

            
    def stream_excel_to_frontend(self, excel_wb):
        """openpyxl传输wb对象到excel数据流"""
        try:
            # 创建一个BytesIO对象来保存Excel文件
            excel_stream = io.BytesIO()
            
            # 将workbook保存到这个BytesIO流中
            excel_wb.save(excel_stream)

            # 重置流的位置到开始处，这样就可以读取它的内容了
            excel_stream.seek(0)

            # 返回流对象
            return excel_stream
        except IOError as e:
            print(f"An error occurred during streaming: {e}")
            return None

    def convert_excel_format(self,input_bytes, src_format, dst_format,save_dst=False):
        """根据参数将数据流中的excel格式进行转化，并输出为数据流,默认在temp文件夹中产生的临时文件"""
        # 确保源格式和目标格式是受支持的
        if src_format not in self.FORMATS or dst_format not in self.FORMATS:
            raise ValueError('Unsupported format specified.')

        src_tempfile_path=os.path.join(self.temp_path,f"temp.{src_format}")
        dst_tempfile_path=os.path.join(self.temp_path,f"temp.{dst_format}")
        
        # 创建 Excel 对象
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False  # 不显示Excel界面
        
        # 创建输出流
        output_io = io.BytesIO()

        # 将输入BytesIO对象中的内容写入临时源文件
        with open(src_tempfile_path, "wb") as temp_file:
            temp_file.write(input_bytes.getvalue())
        print("*********************",os.path.abspath(src_tempfile_path))
        # 打开源文件
        workbook = excel.Workbooks.Open(os.path.abspath(src_tempfile_path))

        # 另存为目标格式的文件
        workbook.SaveAs(dst_tempfile_path, FileFormat=self.FORMATS[dst_format])
        workbook.Close(False)

        # 读取目标文件到BytesIO对象
        with open(dst_tempfile_path, "rb") as temp_file:
            output_io.write(temp_file.read())

        # 清理临时文件
        os.remove(src_tempfile_path)
        
        #在最后的保存excel步骤，可先保留文件至temp文件夹，再传输到用户选择的文件夹
        if not save_dst:
            os.remove(dst_tempfile_path)

        # 关闭 Excel 进程
        excel.Application.Quit()

        # 设置输出流的指针回到起始位置，以便于读取
        output_io.seek(0)
        return output_io


class Excel_attribute:
    """目前只考虑了一个工作簿&其一个工作表的修改，进一步：无法实现多个工作表同时修改"""
    def __init__(self, excel_wb=None , excel_ws=None):
        """类无传输值分别表示创建新wb、读取wb第一个工作表"""
        if excel_wb is None:
            self.excel_wb = px.Workbook()
            self.excel_ws = self.excel_wb.active
        else:
            self.excel_wb = excel_wb
            self.excel_ws = excel_ws if excel_ws is not None else excel_wb.worksheets[0]
    
    def get_some_axis_cells(self,index,value_only=True):
        """获取某一行/列的单元格，依据参数返回单元格对象或值的list"""
        transform_cell=lambda cell:cell.value if value_only==True else cell
        excel_field=[transform_cell(cell) for cell in self.excel_ws[index] if cell.value]
        return excel_field
    
    def get_max_row_col(self):
        """worksheet提供的属性来获取最大行列数问题：目前发现单元格有颜色填充、字色等也会被视为有内容的单元格；
           根据值遍历出的最大行列数则无此问题
           此外，纯下拉列表无选择值，二者都不会视为单元格有内容
           故返回两种方法分别产生的最大行列数集合"""
        px_max_row = self.excel_ws.max_row
        px_max_col = self.excel_ws.max_column
        value_max_row = 0
        value_max_col = 0
        for row in self.excel_ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    value_max_row = max(value_max_row, cell.row)
                    value_max_col = max(value_max_col, cell.column)
        return {"max_row":{px_max_col,value_max_col},
                "max_col":{px_max_row,value_max_row}}

    def modify_cell_style(self, cell, font=None, border=None, fill=None, 
                          number_format=None, protection=None, 
                          hyperlink=None, alignment=None):
        """根据对象性参数修改某一单元格的字体、边框、填充、数字格式、保护方式、超文本、对齐格式等"""
        # Check if cell is a string reference or a Cell object
        if isinstance(cell, str):
            cell = self.excel_ws[cell]
            
        # Apply styles as provided
        # Create a dictionary with attribute names and the values provided
        style_attributes = {
            'font': font,'border': border,'alignment': alignment,
            'fill': fill,'number_format': number_format,
            'protection': protection,'hyperlink': hyperlink }

        # Apply styles as provided
        for attr_name, attr_value in style_attributes.items():
            if attr_value is not None:
                setattr(cell, attr_name, attr_value)  # Set attribute by name
    
    def modify_CertainRange_style(self, cell_range, **style_kwargs):
        """根据对象性参数修改某一单元格区域的字体、边框、填充、数字格式、保护方式、超文本、对齐格式等"""
        # Convert cell range to actual range
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        
        # Iterate over all cells in the range
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = self.excel_ws.cell(row=row, column=col)
                self.modify_cell_style(cell, **style_kwargs)
                
    def modify_MutipleRange_style(self, cell_range_list, **style_kwargs):
        """根据对象性参数修改某一多分布单元格区域的字体、边框、填充、数字格式、保护方式、超文本、对齐格式等"""
        if type(cell_range_list)== str:cell_range_list=[cell_range_list]
        for cell_range in cell_range_list:
            self.modify_CertainRange_style(cell_range,**style_kwargs)

    
    def get_dropdowns(self):
        """获取工作表内的各列的下拉列表字典，同一行多种下拉列表的以list组织"""
        def get_dropdowns_values(validation):
            result=validation.formula1
            
            # 进一步，下拉列表不仅仅为序列
            # 若值为工作表单元格引用
            
            #捕获组 (.*!)? 是可选的，用来匹配任意字符后跟一个感叹号 !，代表可能存在的工作表名称。
            pattern = r"^(.*!)?(\$?[A-Za-z]\$?\d+:\$?[A-Za-z]\$?\d+)$"
            match_=re.search(pattern,result)
            if match_:
                match_groups=match_.groups()
                # 若跨工作表引用(预计更为合理)
                if (match_groups)[0]:
                    sheet_name=match_groups[0][:-1]
                    sheet_name=sheet_name[1:-1] if sheet_name[0]==sheet_name[-1]=="'" else sheet_name
                    dropdown_sourcesheet=self.excel_wb[sheet_name]
                else:dropdown_sourcesheet=self.excel_ws
                # 默认被引用为数据验证的单元格不止一个
                min_col, min_row, max_col, max_row = range_boundaries(match_groups[-1].replace('$', ''))
                value_list=[]
                for i in range(min_row, max_row+1):
                    for j in range(min_col, max_col+1):
                        value_list.append(dropdown_sourcesheet.cell(i, j).value)
                return value_list
            
            # 若值为简单的手动输入序列
            elif "," in result:
                # 去除首尾的引号后，直接拆分为值
                return result[1:-1].split(',')
            
        drop_row=dict()

        # 含有当前工作表的所有有效性验证的对象
        validations = self.excel_ws.data_validations.dataValidation
        print(validations)
        for validation in validations:
            
            #当前有效性涉及区域
            cell=str(validation.sqref)
            
            #目前的方式，仅匹配下拉列表选择所以值的。进一步：考虑介于等多种方式
            result=(get_dropdowns_values(validation))

            #如果是多列的下拉列表相同，分别进行检验
            if " " in cell:
                cells=cell.split(" ")
                for i in cells:
                    if i[0] not in drop_row:drop_row[i[0]]=[result]
                    elif set(result) in [set(already_result) for already_result in drop_row[i[0]]]:continue
                    else:drop_row[i[0]].append(result)
            else:
                if (cell)[0] not in drop_row:drop_row[(cell)[0]]=[result]
                else:drop_row[cell[0]].append(result)
        return drop_row
         
    def create_or_update_dv_list(self, field, rule_list):
        """将过长的下拉列表中的选项写入隐藏的工作表中,工作表中各列内容：
            A           B           C
        1   A1序号      C1院系(新表列号与旧表列号不必一直)
        2   下拉列表值1  下拉列表值1  .
        3   下拉列表值2  下拉列表值2  .
        4   下拉列表值3  下拉列表值3  .
        5   下拉列表值4  下拉列表值4  .
        6   下拉列表值5  下拉列表值5  .
        7   下拉列表值6  下拉列表值6  .
        """
        # 检查是否存在名为'sheet_for_DataValidate'的工作表，如果没有则创建
        sheet_name = 'sheet_for_DataValidate'
        if sheet_name not in self.excel_wb.sheetnames:
            dv_sheet = self.excel_wb.create_sheet(sheet_name)
            dv_sheet.sheet_state = 'hidden'  # 隐藏工作表
        else:
            dv_sheet = self.excel_wb[sheet_name]

        # 查找为空的列（即没有数据验证列表的列）
        col_index = 1
        while dv_sheet.cell(row=2, column=col_index).value is not None:
            col_index += 1

        # 在找到的列的第一行中写入字段名
        dv_sheet.cell(row=1, column=col_index, value=field)

        # 从第二行开始写入规则列表
        for index, item in enumerate(rule_list, start=2):
            dv_sheet.cell(row=index, column=col_index, value=item)

        # 返回引用区域的字符串，例如'Sheet2!$B$2:$B$100'
        return f"'{sheet_name}'!${dv_sheet.cell(row=2, column=col_index).column_letter}$2:${dv_sheet.cell(row=2, column=col_index).column_letter}${len(rule_list)+1}"


    def set_dropdowns(self, selected_field_rules, sep_row=2):
        """将用户选定的规则字典导出到下拉列表，默认假设字段在第n行，在第n+1行添加规则样例行，n+2开始是下拉列表"""
        selected_field_rules = {k: v for k, v in selected_field_rules.items() if v[1]}  # 去掉规则列表没有内容的字段
        
        for one_index_col, (field_name, rule_list) in selected_field_rules.items():
            if rule_list:  # 当规则列表有内容时
                dv_col, dv_beginrow = one_index_col[0], int(one_index_col[1:]) + sep_row
                sqref = f'{dv_col}{dv_beginrow}:{dv_col}1048576'  # 确保范围引用格式正确
                
                # 将规则列表转化为逗号分隔的字符串,并检测是否比255长，若过长即采用引用区域方式呈现。
                formula1_insides_quotes = ",".join(rule_list)
                if len(formula1_insides_quotes)>250:
                    formula1=self.create_or_update_dv_list(one_index_col+":"+field_name,rule_list)
                else:formula1=f'"{formula1_insides_quotes}"'
                # 添加下拉列表及其对应区域
                dv = DataValidation(type="list", formula1=formula1, showErrorMessage=True, allow_blank=False)
                self.excel_ws.add_data_validation(dv)
                dv.add(sqref)
                
                
def convert_to_json_stream(data):
    """将Python数据类型转化为JSON格式的字符串，后端不再使用。"""
    json_string = json.dumps(data)
    
    # 创建一个StringIO对象，它提供了文件类的接口
    json_stream = io.StringIO(json_string)
    
    # 返回数据流
    return json_stream

def read_from_json_stream(json_stream):
    """从JSON数据流中读取数据并转换为Python数据类型，后端不再使用。"""
    # 重置流的读取位置到起始处
    json_stream.seek(0)
    
    # 从数据流中读取JSON数据并转换为Python数据类型
    data = json.load(json_stream)
    
    # 返回Python数据类型
    return data
def read_from_json_file(file_path):
    """从JSON文件中读取数据并转换为Python数据类型"""
    with open(file_path, 'r',encoding="utf-8") as json_file:
        data = json.load(json_file)
    return data

if "__main__" == __name__:
    """
    excel_got="tests/for_fuker.func2/test2_dropdown.xlsx"
    print(os.listdir())
    Xio=Excel_IO()
    xlsx_wb=px.load_workbook(excel_got)
    xlsx_stream=Xio.stream_excel_to_frontend(xlsx_wb)
    Xattr=Excel_attribute(xlsx_wb)
    Xattr.get_dropdowns()
    #Xio.convert_excel_format(xlsx_stream,"xlsx","xls",True)
    """
    # 测试将规则写入下拉列表
    selected_field_rules={'a5': ('序号', []), 'b5': ('作品题目', []), 'c5': ('参赛类别', []), 'd5': ('作品学科分类', ['理工农医类', '社会调查报告和人文社科类', '发明创造科技制作类']), 'e5': ('学科门类', ['哲学', '经济学', '法学', '教育学', '文学', '历史学', '理学', '工学', '农学', '医学', '军事学', '管理学', '艺术学']), 'f5': ('一级学科', ['哲学', '理论经济学', '应用经济学', '法学', '政治学', '社会学', '民族学', '马克思主义理论', '公安学', '教育学', '心理学', '体育学', '中国语言文学', '外国语言文学', '新闻传播学', '考古学', '中国史', '世界史', '数学', '物理学', '化学', '天文学', '地理学', '大气科学', '海洋科学', '地球物理学', '地质学', '生物学', '系统科学', '科学技术史', '生态学', '统计学', '力学', '机械工程', '光学工程', '仪器科学与技术', '材料科学与工程', '冶金工程', '动力工程及工程热物理', '电气工程', '电子科学与技术', '信息与通信工程', '控制科学与工程', '计算机科学与技术', '建筑学', '土木工程', '水利工程', '测绘科学与技术', '化学工程与技术', '地质资源与地质工程', '矿业工程', '石油与天然气工程', '纺织科学与工程', '轻工技术与工程', '交通运输工程', '船舶与海洋工程', '航空宇航科学与技术', '兵器科学与技术', '核科学与技术', '农业工程', '林业工程', '环境科学与工程', '生物医学工程', '食品科学与工程', '城乡规划学', '风景园林学', '软件工程', '生物工程', '安全科学与工程', '公安技术', '网络空间安全', '作物学', '园艺学', '农业资源与环境', '植物保护', '畜牧学', '兽医学', '林学', '水产', '草学', '基础医学', '临床医学', '口腔医学', '公共卫生与预防医学', '中医学', '中西医结合', '药学', '中药学', '特种医学', '医学技术', '护理学', '军事思想及军事历史', '战略学', '战役学', '战术学', '军队指挥学', '军事管理学', '军队政治工作学', '军事后勤学', '军事装备学', '军事训练学', '管理科学与工程', '工商管理', '农林经济管理', '公共管理', '图书情报与档案管理', '艺术学理论', '音乐与舞蹈学', '戏剧与影视学', '美术学', '设计学']), 'g5': ('作者', []), 'h5': ('是否为团队负责人', ['是', '否']), 'i5': ('性别', ['男', '女']), 'j5': ('生源地', ['北京', '天津', '上海', '重庆', '河北', '山西', '辽宁', '吉林', '黑龙江', '江苏', '浙江', '安徽', '福建', '江西', '山东', '河南', '湖北', '湖南', '广东', '海南', '四川', '贵州', '云南', '陕西', '甘肃', '青海', '台湾', '内蒙古', '广西', '西藏', '宁夏', '新疆', '香港', '澳门', '其他']), 'k5': ('学号', []), 'l5': ('所在院系', ['数学科学学院', '物理学院', '化学与分子工程学院', '生命科学学院', '地球与空间科学学院', '城市与环境学院', '心理与认知科学学院', '建筑与景观设计学院', '信息科学技术学院', '工学院', '王选计算机研究所', '软件与微电子学院', '环境科学与工程学院', '软件工程国家工程研究中心', '中国语言文学系', '历史学系', '考古文博学院', '哲学系', '外国语学院', '艺术学院', '对外汉语教育学院', '歌剧研究院', '国际关系学院', '法学院', '信息管理系', '社会学系', '政府管理学院', '马克思主义学院', '教育学院', '新闻与传播学院', '体育教研部', '新媒体研究院', '教育财政科学研究所', '经济学院', '光华管理学院', '人口研究所', '国家发展研究院', '基础医学院', '药学院', '公共卫生学院', '护理学院', '医学人文学院', '医学继续教育学院', '第一医院', '人民医院', '第三医院', '口腔医院', '北京肿瘤医院', '第六医院', '深圳医院', '首钢医院', '国际医院', '滨海医院', '元培学院', '燕京学堂', '先进技术研究院', '前沿交叉学科研究院', '中国社会科学调查中心', '分子医学研究所', '科维理天文研究所', '核科学与技术研究院', '北京国际数学研究中心', '海洋研究院', '现代农学院', '人文社会科学研究院', '信息工程学院', '化学生物学与生物技术学院', '环境与能源学院', '城市规划与设计学院', '新材料学院', '汇丰商学院', '国际法学院', '人文社会科学学院']), 'm5': ('年级（如2020级本科生/硕士生/博士生）', []), 'n5': ('手机', []), 'o5': ('微信号', []), 'p5': ('邮箱', []), 'q5': ('指导教师姓名', []), 'r5': ('指导教师性别', ['男', '女']), 's5': ('指导教师所在院系', ['数学科学学院', '物理学院', '化学与分子工程学院', '生命科学学院', '地球与空间科学学院', '城市与环境学院', '心理与认知科学学院', '建筑与景观设计学院', '信息科学技术学院', '工学院', '王选计算机研究所', '软件与微电子学院', '环境科学与工程学院', '软件工程国家工程研究中心', '中国语言文学系', '历史学系', '考古文博学院', '哲学系', '外国语学院', '艺术学院', '对外汉语教育学院', '歌剧研究院', '国际关系学院', '法学院', '信息管理系', '社会学系', '政府管理学院', '马克思主义学院', '教育学院', '新闻与传播学院', '体育教研部', '新媒体研究院', '教育财政科学研究所', '经济学院', '光华管理学院', '人口研究所', '国家发展研究院', '基础医学院', '药学院', '公共卫生学院', '护理学院', '医学人文学院', '医学继续教育学院', '第一医院', '人民医院', '第三医院', '口腔医院', '北京肿瘤医院', '第六医院', '深圳医院', '首钢医院', '国际医院', '滨海医院', '元培学院', '燕京学堂', '先进技术研究院', '前沿交叉学科研究院', '中国社会科学调查中心', '分子医学研究所', '科维理天文研究所', '核科学与技术研究院', '北京国际数学研究中心', '海洋研究院', '现代农学院', '人文社会科学研究院', '信息工程学院', '化学生物学与生物技术学院', '环境与能源学院', '城市规划与设计学院', '新材料学院', '汇丰商学院', '国际法学院', '人文社会科学学院']), 't5': ('指导教师职称/职务', []), 'u5': ('指导教师电话', []), 'v5': ('指导教师电子邮箱', [])}

    excel_got="tests/for_fuker.func2/test_set_dropdown.xlsx"
    Xio=Excel_IO()
    xlsx_wb=px.load_workbook(excel_got)
    Xattr=Excel_attribute(xlsx_wb)
    Xattr.set_dropdowns(selected_field_rules)
    print(Xattr.get_dropdowns())
    xlsx_wb.save("tests/for_fuker.func2/output2-test_set_dropdown.xlsx")
    #"""